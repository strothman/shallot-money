import os
import json
import datetime
import openpyxl

excel_path = r"\\joandesk\Cloud\-Rick\Rick Spreadsheet Budget.xlsx"
output_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "src", "imported_expenses.json")

print(f"Reading spreadsheet from: {excel_path}")
if not os.path.exists(excel_path):
    print("Error: Excel budget file not found.")
    exit(1)

wb = openpyxl.load_workbook(excel_path, read_only=True)
if 'Expenses' not in wb.sheetnames:
    print("Error: 'Expenses' sheet not found in the Excel workbook.")
    exit(1)

sheet = wb['Expenses']
expenses = []

# Excel category mapping to App category IDs
category_mapping = {
    'Food': 'food',
    'Groceries': 'food',
    'Gas': 'transport',
    'Gym': 'gym',
    'Prime': 'bills',
    'Entertainment': 'entertainment',
    'Merch': 'shopping',
    'Goods?': 'shopping',
    'Fast Food': 'fastfood'
}

for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
    # Skip header
    if r_idx == 0:
        continue
        
    # Check if we have enough columns and the row is not empty
    if len(row) < 5:
        continue
        
    date_val = row[1]
    business = row[2]
    amount = row[3]
    category = row[4]
    notes = row[5] if len(row) > 5 else None

    # Skip empty lines
    if date_val is None or amount is None:
        continue

    # Date Parsing
    date_str = ""
    if isinstance(date_val, (datetime.datetime, datetime.date)):
        date_str = date_val.strftime('%Y-%m-%d')
    else:
        # String parsing
        date_val = str(date_val).strip()
        for fmt in ('%m-%d-%Y', '%m/%d/%Y', '%Y-%m-%d'):
            try:
                dt = datetime.datetime.strptime(date_val, fmt)
                date_str = dt.strftime('%Y-%m-%d')
                break
            except ValueError:
                pass
        if not date_str:
            # Fallback to today if unparseable
            date_str = datetime.date.today().strftime('%Y-%m-%d')

    # Description mapping (combine Business and Notes if notes exists)
    raw_biz = str(business or "Unknown").strip()
    notes_str = str(notes).strip() if notes else ""
    desc_str = f"{raw_biz} ({notes_str})" if notes_str else raw_biz
    items_list = [it.strip() for it in notes_str.split(',') if it.strip()] if notes_str else []

    # Category Mapping
    cat_str = str(category).strip()
    mapped_category = category_mapping.get(cat_str, 'gym')

    # Amount Mapping
    try:
        amt_float = float(amount)
    except ValueError:
        continue

    # Create expense object (Generate stable IDs from row number and values)
    exp_entry = {
        "id": f"imported_{r_idx}_{date_str}",
        "amount": amt_float,
        "description": desc_str,
        "category": mapped_category,
        "date": date_str
    }
    if items_list:
        exp_entry["items"] = items_list
    expenses.append(exp_entry)

# Write to src/imported_expenses.json
os.makedirs(os.path.dirname(output_path), exist_ok=True)
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(expenses, f, indent=2, ensure_ascii=False)

print(f"Successfully sync'd {len(expenses)} expenses to {output_path}")
